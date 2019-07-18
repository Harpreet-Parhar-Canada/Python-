# Classes to construct Client, Invoice, and Product Objects for use in Helper class:
class Clients():
    def __init__(self, name):
        self.name = name

class Product():
    def __init__(self, name, price, inventory):
        self.name = name
        self.price = price
        self.inventory = inventory

class Invoice():
    def __init__(self, clientid, date):
        self.clientid = clientid
        self.date = date
        self.product = ''

# Helper class: makes dictionaries from excel sheets and performs 'worker' functions on excel data
class Helper():
    def __init__(self, book):
        self.book = book
        from openpyxl import load_workbook, Workbook
        self.wb = load_workbook(book)
        self.cust_ws = self.wb['Clients']
        self.invoice_ws = self.wb['Invoices']
        self.itsold_ws = self.wb['Items Sold']
        self.product_ws = self.wb['Product Inventory']

    def make_cust_dict(self):
        self.client_dict = {}
        self.dup_list = []
        for row in self.cust_ws.iter_rows(min_row=2, values_only=True):
            if row[0] not in self.client_dict and row[0] != None and row[1] != None:
                self.client_dict[row[0]] = Clients(row[1])
            else:
                self.dup_list.append({row[0]:row[1]})
        print('Duplicate Client Data', self.dup_list)

        return self.client_dict

    def make_prod_dict(self):
        self.prod_dict = {}
        self.dupprod_list = []
        for row in self.product_ws.iter_rows(min_row=2, values_only=True):
            if row[0] not in self.prod_dict and row[0] != None and row[1] != None and row[2] != None and row[3] != None:
                self.prod_dict[row[0]] = Product(row[1], row[2], row[3])
            else:
                self.dupprod_list.append({row[0]:row[1]})
        print('Duplicate Product Data', self.dupprod_list)

        return self.prod_dict

    def item(self, invoiceid):
        self.item_dict = {}
        self.missing_list = []
        for row in self.itsold_ws.iter_rows(min_row=2, values_only=True):
            if row[0] != None and row[1] != None and row[2] != None:
                if row[0] == invoiceid:
                    self.item_dict[row[1]] = (row[2])
            else:
                self.missing_list.append({row[0]: {row[1]: row[2]}})

        return self.item_dict

    def make_inv_dict(self):
        self.inv_dict = {}
        self.dupinv_list = []
        for row in self.invoice_ws.iter_rows(min_row=2, values_only=True):
            if row[0] not in self.inv_dict and row[1] != None and row[2] != None:
                self.inv_dict[row[0]] = Invoice(row[1], row[2])
                self.inv_dict[row[0]].product = self.item(row[0])
            else:
                self.dupinv_list.append({row[0]:row[1]})
        print("Duplicate Invoice Data", self.dupinv_list)

        return self.inv_dict

    def make_invoice(self, invoiceid):
        from openpyxl import load_workbook
        invoice_output = load_workbook('Invoice_template.xlsx')
        ws = invoice_output['Invoice']

        self.make_cust_dict()
        self.make_prod_dict()
        self.make_inv_dict()

        b = self.inv_dict[invoiceid].clientid

        ws['B3'] = invoiceid
        ws['B4'] = self.inv_dict[invoiceid].date
        ws['B5'] = self.client_dict[b].name
        ws['B6'] = self.inv_dict[invoiceid].clientid

        c = self.inv_dict[invoiceid].product
        prod_list = []
        count = 0
        inv_total = 0
        for i in c:
            pd_name = self.prod_dict[i].name
            pd_id = i
            pd_price = self.prod_dict[i].price
            pd_quantity = c[i]
            total_pd_cost = pd_price * pd_quantity
            inv_total = inv_total + total_pd_cost
            prod_list.extend([[pd_name, pd_id, pd_price, pd_quantity, total_pd_cost]])
            count = count + 1

        for i, line in enumerate(prod_list):
            ws.cell(row=i+9, column=1, value=line[0])
            ws.cell(row=i+9, column=2, value=line[1])
            ws.cell(row=i+9, column=3, value=line[2])
            ws.cell(row=i+9, column=4, value=line[3])
            ws.cell(row=i+9, column=5, value=line[4])
        ws[f'D{10 + count}'] = 'Total'
        ws[f'E{10 + count}'] = inv_total

        newfilename = f"invoice {str(invoiceid)}.xlsx"
        invoice_output.save(newfilename)

new_obj = Helper('Invoice-Example.xlsx')
new_obj.make_invoice('00015')




