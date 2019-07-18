from openpyxl import load_workbook

wb = load_workbook("Invoice-Example.xlsx")

class Basic_obj(dict):

   def __getattr__(self, item):
       return self[item]


def workSheets():
    return wb.sheetnames

def make_object_dictionary_value(props_name, obj_values):
    new_obj = Basic_obj()
    for i in range(len(props_name)):
        new_obj[props_name[i]] = obj_values[i]
    return (new_obj)

def display_dictionary_line_by_line(dict):
    print(' ')
    for item in dict:
        print(item, ':', dict[item])

def make_dictionaries_for_each_sheet(entire_sheet):
    temp_dict = {}
    counter = 0
    objects_property_names = ""
    for row in entire_sheet.values:
        if counter == 0:
            objects_property_names = list(row)
            objects_property_names.pop(0)
            counter += 1

        else:
            object_property_values = list(row)
            the_key = object_property_values.pop(0)
            the_value = make_object_dictionary_value(objects_property_names, object_property_values)

def read_excel_file():
    master_wb = Basic_obj()
    for sheet in wb:
        ws = wb.get_sheet_by_name(sheet.title)
        master_wb_value = make_dictionaries_for_each_sheet(ws)
        master_wb[sheet.title] = master_wb_value
        print(master_wb_value)
        print(len(master_wb_value))
        # display_dictionary_line_by_line(master_wb_value)





print(workSheets())
read_excel_file()

